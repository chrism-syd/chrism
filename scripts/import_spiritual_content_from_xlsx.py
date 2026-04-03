#!/usr/bin/env python3
"""
Import spiritual companion workbook content into the normalized Chrism schema.

Usage:
  DATABASE_URL=postgresql://... python scripts/import_spiritual_content_from_xlsx.py \
    --xlsx /path/to/saints_master_dataset_v1.xlsx \
    --group-tag-map scripts/spiritual_group_tag_scope.example.json \
    --write

By default, the script runs in DRY RUN mode and rolls back at the end.
Pass --write to commit.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import psycopg
from psycopg.rows import dict_row


MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

TOPIC_SPLIT_RE = re.compile(r"[|,;]")
WHITESPACE_RE = re.compile(r"\s+")
SLUG_RE = re.compile(r"[^a-z0-9]+")

WORKBOOK_SHEET_NAMES = [
    "saints",
    "patronage_categories",
    "topic_tags",
    "saint_topic_map",
    "scripture_passages",
    "prayers",
    "catechism_references",
]


@dataclass
class ImportStats:
    topics_upserted: int = 0
    topic_aliases_upserted: int = 0
    saints_upserted: int = 0
    saint_aliases_upserted: int = 0
    saint_topics_upserted: int = 0
    scripture_upserted: int = 0
    scripture_topics_upserted: int = 0
    catechism_upserted: int = 0
    catechism_topics_upserted: int = 0
    prayers_upserted: int = 0
    prayer_topics_upserted: int = 0
    prayer_scopes_replaced: int = 0
    skipped_patronage_categories: int = 0
    warnings: List[str] = field(default_factory=list)

    def warn(self, message: str) -> None:
        if message not in self.warnings:
            self.warnings.append(message)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return WHITESPACE_RE.sub(" ", str(value)).strip()


def normalize_bool(value: Any) -> bool:
    return normalize_text(value).lower() in {"true", "1", "yes", "y"}


def slugify(value: str) -> str:
    cleaned = SLUG_RE.sub("-", normalize_text(value).lower()).strip("-")
    return cleaned or "untitled"


def titleize_slug(value: str) -> str:
    return " ".join(part.capitalize() for part in slugify(value).split("-"))


def split_multi_value(value: Any) -> List[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    parts = [normalize_text(part) for part in TOPIC_SPLIT_RE.split(raw)]
    return [part for part in parts if part]


def parse_date(value: Any) -> Optional[datetime]:
    raw = normalize_text(value)
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def parse_feast_day(value: Any) -> Tuple[Optional[int], Optional[int]]:
    raw = normalize_text(value)
    if not raw:
        return None, None
    match = re.match(r"^\s*([A-Za-z]+)\s+(\d{1,2})\s*$", raw)
    if not match:
        return None, None
    month = MONTHS.get(match.group(1).lower())
    day = int(match.group(2))
    return month, day if month else (None, None)


def read_sheet_rows(workbook_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing worksheet: {sheet_name}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [normalize_text(cell) for cell in rows[0]]
        data: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if not row or not any(cell is not None and normalize_text(cell) for cell in row):
                continue
            data.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
        return data
    finally:
        wb.close()


def map_relevance(value: Any) -> Optional[int]:
    raw = normalize_text(value).lower()
    if not raw:
        return None
    if raw == "primary":
        return 5
    if raw == "secondary":
        return 3
    if raw == "tertiary":
        return 1
    return None


def map_prayer_type(value: Any) -> Optional[str]:
    raw = normalize_text(value).lower()
    mapping = {
        "common_prayer": "traditional",
        "creed": "devotion",
        "doxology": "traditional",
        "marian": "devotion",
        "intercessory": "intercession",
        "franciscan": "devotion",
        "devotional_grouping": "devotion",
        "rosary_grouping": "devotion",
        "discernment": "devotion",
        "petition": "intercession",
        "litany": "litany",
        "novena": "novena",
        "chaplet": "chaplet",
        "blessing": "blessing",
        "collect": "collect",
    }
    return mapping.get(raw, "other" if raw else None)


def map_text_status(value: Any) -> str:
    raw = normalize_text(value).lower()
    if raw in {"normative", "official_devotional_text", "official_group_prayer"}:
        return "published"
    if raw in {"custom_user_text"}:
        return "draft"
    if raw in {"published"}:
        return "published"
    if raw in {"approved"}:
        return "approved"
    if raw in {"review"}:
        return "review"
    if raw in {"retired", "inactive"}:
        return "retired"
    return "draft"


class SpiritualImporter:
    def __init__(self, conn: psycopg.Connection, workbook_path: str, dry_run: bool, group_tag_map: Optional[Dict[str, Any]], allow_unmapped_group_content_as_global: bool):
        self.conn = conn
        self.workbook_path = workbook_path
        self.dry_run = dry_run
        self.group_tag_map = group_tag_map or {}
        self.allow_unmapped_group_content_as_global = allow_unmapped_group_content_as_global
        self.stats = ImportStats()

        self.topic_id_by_workbook_id: Dict[str, str] = {}
        self.topic_id_by_key: Dict[str, str] = {}
        self.saint_id_by_workbook_id: Dict[str, str] = {}
        self.saint_id_by_name: Dict[str, str] = {}
        self.org_family_id_by_code: Dict[str, str] = {}
        self.local_unit_id_by_code: Dict[str, str] = {}

    def run(self) -> ImportStats:
        self._load_scope_lookup_tables()
        self._import_topics()
        self._import_saints()
        self._import_saint_topics()
        self._import_scripture()
        self._import_catechism()
        self._import_prayers()
        self.stats.skipped_patronage_categories = len(read_sheet_rows(self.workbook_path, "patronage_categories"))
        self.stats.warn("The patronage_categories worksheet is currently not imported because the starter schema does not yet include patronage tables.")
        return self.stats

    def _load_scope_lookup_tables(self) -> None:
        try:
            with self.conn.cursor(row_factory=dict_row) as cur:
                rows = cur.execute("select id, code from public.organization_families").fetchall()
                self.org_family_id_by_code = {
                    normalize_text(row["code"]).lower(): str(row["id"]) for row in rows if normalize_text(row.get("code"))
                }
        except Exception as exc:
            self.conn.rollback()
            self.stats.warn(
                f"Could not load organization_families(code): {exc}. "
                "Group-specific prayers will need explicit scope mapping or will remain unmapped."
            )

        try:
            with self.conn.cursor(row_factory=dict_row) as cur:
                rows = cur.execute("select id, code from public.local_units").fetchall()
                self.local_unit_id_by_code = {
                    normalize_text(row["code"]).lower(): str(row["id"]) for row in rows if normalize_text(row.get("code"))
                }
        except Exception as exc:
            self.conn.rollback()
            self.stats.warn(
                f"Could not load local_units(code): {exc}. "
                "Local-unit spiritual scope mapping will need explicit IDs if used."
            )

    def _upsert_topic(self, *, slug: str, name: str, topic_group: Optional[str], description: Optional[str], sort_order: int = 0) -> str:
        with self.conn.cursor() as cur:
            topic_id = cur.execute(
                """
                insert into public.spiritual_topics (
                  slug, name, topic_group, description, is_active, sort_order
                )
                values (%s, %s, %s, %s, true, %s)
                on conflict (slug) do update
                  set name = excluded.name,
                      topic_group = excluded.topic_group,
                      description = coalesce(excluded.description, public.spiritual_topics.description),
                      is_active = true,
                      sort_order = excluded.sort_order,
                      updated_at = timezone('utc', now())
                returning id
                """,
                (slug, name, topic_group, description, sort_order),
            ).fetchone()["id"]
        self.stats.topics_upserted += 1
        return str(topic_id)

    def _insert_topic_alias(self, topic_id: str, alias: str) -> None:
        alias = normalize_text(alias)
        if not alias:
            return
        with self.conn.cursor() as cur:
            cur.execute(
                """
                insert into public.spiritual_topic_aliases (topic_id, alias)
                values (%s, %s)
                on conflict do nothing
                """,
                (topic_id, alias),
            )
        self.stats.topic_aliases_upserted += 1

    def _register_topic_keys(self, topic_id: str, workbook_id: Optional[str], name: str, aliases: Sequence[str]) -> None:
        if workbook_id:
            self.topic_id_by_workbook_id[normalize_text(workbook_id)] = topic_id
        self.topic_id_by_key[normalize_text(name).lower()] = topic_id
        self.topic_id_by_key[slugify(name)] = topic_id
        for alias in aliases:
            normalized = normalize_text(alias)
            if normalized:
                self.topic_id_by_key[normalized.lower()] = topic_id
                self.topic_id_by_key[slugify(normalized)] = topic_id

    def _resolve_topic(self, token: str) -> str:
        key = normalize_text(token)
        if not key:
            raise ValueError("Cannot resolve empty topic token.")
        cached = self.topic_id_by_workbook_id.get(key) or self.topic_id_by_key.get(key.lower()) or self.topic_id_by_key.get(slugify(key))
        if cached:
            return cached

        # Create a lightweight dynamic topic when the workbook references something not in topic_tags.
        topic_id = self._upsert_topic(
            slug=slugify(key),
            name=titleize_slug(key),
            topic_group="Imported",
            description=None,
        )
        self._register_topic_keys(topic_id, None, titleize_slug(key), [])
        self.stats.warn(f"Created dynamic topic for token '{key}' because it was not found in topic_tags.")
        return topic_id

    def _import_topics(self) -> None:
        rows = read_sheet_rows(self.workbook_path, "topic_tags")
        for index, row in enumerate(rows, start=1):
            topic_name = normalize_text(row.get("Topic Name"))
            if not topic_name:
                continue
            aliases = split_multi_value(row.get("Topic Aliases"))
            topic_id = self._upsert_topic(
                slug=slugify(topic_name),
                name=topic_name,
                topic_group=normalize_text(row.get("Category")) or None,
                description=normalize_text(row.get("Description")) or None,
                sort_order=index * 10,
            )
            for alias in aliases:
                self._insert_topic_alias(topic_id, alias)
            self._register_topic_keys(topic_id, normalize_text(row.get("Topic ID")) or None, topic_name, aliases)

    def _upsert_saint(self, *, slug: str, canonical_name: str, common_name: Optional[str], short_bio: Optional[str], feast_month: Optional[int], feast_day: Optional[int], era_label: Optional[str], canonization_status: Optional[str], patron_summary: Optional[str]) -> str:
        with self.conn.cursor() as cur:
            saint_id = cur.execute(
                """
                insert into public.saints (
                  slug, canonical_name, common_name, short_bio, feast_month, feast_day,
                  era_label, canonization_status, patron_summary, is_active
                )
                values (%s, %s, %s, %s, %s, %s, %s, %s, %s, true)
                on conflict (slug) do update
                  set canonical_name = excluded.canonical_name,
                      common_name = coalesce(excluded.common_name, public.saints.common_name),
                      short_bio = coalesce(excluded.short_bio, public.saints.short_bio),
                      feast_month = coalesce(excluded.feast_month, public.saints.feast_month),
                      feast_day = coalesce(excluded.feast_day, public.saints.feast_day),
                      era_label = coalesce(excluded.era_label, public.saints.era_label),
                      canonization_status = coalesce(excluded.canonization_status, public.saints.canonization_status),
                      patron_summary = coalesce(excluded.patron_summary, public.saints.patron_summary),
                      is_active = true,
                      updated_at = timezone('utc', now())
                returning id
                """,
                (slug, canonical_name, common_name, short_bio, feast_month, feast_day, era_label, canonization_status, patron_summary),
            ).fetchone()["id"]
        self.stats.saints_upserted += 1
        return str(saint_id)

    def _insert_saint_alias(self, saint_id: str, alias: str) -> None:
        alias = normalize_text(alias)
        if not alias:
            return
        with self.conn.cursor() as cur:
            cur.execute(
                """
                insert into public.saint_aliases (saint_id, alias)
                values (%s, %s)
                on conflict do nothing
                """,
                (saint_id, alias),
            )
        self.stats.saint_aliases_upserted += 1

    def _import_saints(self) -> None:
        rows = read_sheet_rows(self.workbook_path, "saints")
        for row in rows:
            canonical_name = normalize_text(row.get("Canonical Name"))
            if not canonical_name:
                continue

            feast_month, feast_day = parse_feast_day(row.get("Feast Day"))
            alt_names = split_multi_value(row.get("Alternate Names"))
            saint_id = self._upsert_saint(
                slug=slugify(canonical_name),
                canonical_name=canonical_name,
                common_name=normalize_text(row.get("Common Name")) or None,
                short_bio=normalize_text(row.get("Short Bio")) or None,
                feast_month=feast_month,
                feast_day=feast_day,
                era_label=normalize_text(row.get("Era")) or None,
                canonization_status=normalize_text(row.get("Canonization Status")) or None,
                patron_summary=normalize_text(row.get("Patronages")) or None,
            )
            for alias in alt_names:
                self._insert_saint_alias(saint_id, alias)
            workbook_id = normalize_text(row.get("Saint ID"))
            if workbook_id:
                self.saint_id_by_workbook_id[workbook_id] = saint_id
            self.saint_id_by_name[canonical_name.lower()] = saint_id

    def _import_saint_topics(self) -> None:
        rows = read_sheet_rows(self.workbook_path, "saint_topic_map")
        for row in rows:
            saint_workbook_id = normalize_text(row.get("Saint ID"))
            saint_name = normalize_text(row.get("Canonical Name"))
            saint_id = self.saint_id_by_workbook_id.get(saint_workbook_id) or self.saint_id_by_name.get(saint_name.lower())
            if not saint_id:
                self.stats.warn(f"Skipped saint_topic_map row for saint '{saint_name or saint_workbook_id}' because no saint could be resolved.")
                continue
            topic_workbook_id = normalize_text(row.get("Topic ID"))
            topic_name = normalize_text(row.get("Topic Name"))
            topic_id = self.topic_id_by_workbook_id.get(topic_workbook_id) or self.topic_id_by_key.get(topic_name.lower()) or self._resolve_topic(topic_name or topic_workbook_id)
            with self.conn.cursor() as cur:
                cur.execute(
                    """
                    insert into public.saint_topics (
                      saint_id, topic_id, relevance_score, notes
                    )
                    values (%s, %s, %s, %s)
                    on conflict (saint_id, topic_id) do update
                      set relevance_score = excluded.relevance_score,
                          notes = coalesce(excluded.notes, public.saint_topics.notes)
                    """,
                    (saint_id, topic_id, map_relevance(row.get("Relevance")), normalize_text(row.get("Notes")) or None),
                )
            self.stats.saint_topics_upserted += 1

    def _upsert_scripture(self, slug: str, book: str, reference_label: str, summary: Optional[str]) -> str:
        with self.conn.cursor() as cur:
            scripture_id = cur.execute(
                """
                insert into public.scripture_passages (
                  slug, book, reference_label, summary, is_active
                )
                values (%s, %s, %s, %s, true)
                on conflict (slug) do update
                  set book = excluded.book,
                      reference_label = excluded.reference_label,
                      summary = coalesce(excluded.summary, public.scripture_passages.summary),
                      is_active = true,
                      updated_at = timezone('utc', now())
                returning id
                """,
                (slug, book, reference_label, summary),
            ).fetchone()["id"]
        self.stats.scripture_upserted += 1
        return str(scripture_id)

    def _import_scripture(self) -> None:
        rows = read_sheet_rows(self.workbook_path, "scripture_passages")
        for row in rows:
            scripture_id = self._upsert_scripture(
                slug=slugify(normalize_text(row.get("Scripture ID")) or normalize_text(row.get("Reference"))),
                book=normalize_text(row.get("Book")),
                reference_label=normalize_text(row.get("Reference")),
                summary=normalize_text(row.get("Summary")) or None,
            )
            topic_tokens = split_multi_value(row.get("Topics"))
            for token in topic_tokens:
                topic_id = self._resolve_topic(token)
                with self.conn.cursor() as cur:
                    cur.execute(
                        """
                        insert into public.scripture_topics (
                          scripture_passage_id, topic_id, relevance_score
                        )
                        values (%s, %s, %s)
                        on conflict (scripture_passage_id, topic_id) do nothing
                        """,
                        (scripture_id, topic_id, 5),
                    )
                self.stats.scripture_topics_upserted += 1

    def _upsert_catechism(self, slug: str, reference_code: str, title: Optional[str], summary: Optional[str]) -> str:
        with self.conn.cursor() as cur:
            catechism_id = cur.execute(
                """
                insert into public.catechism_references (
                  slug, reference_code, title, summary, is_active
                )
                values (%s, %s, %s, %s, true)
                on conflict (slug) do update
                  set reference_code = excluded.reference_code,
                      title = coalesce(excluded.title, public.catechism_references.title),
                      summary = coalesce(excluded.summary, public.catechism_references.summary),
                      is_active = true,
                      updated_at = timezone('utc', now())
                returning id
                """,
                (slug, reference_code, title, summary),
            ).fetchone()["id"]
        self.stats.catechism_upserted += 1
        return str(catechism_id)

    def _import_catechism(self) -> None:
        rows = read_sheet_rows(self.workbook_path, "catechism_references")
        for row in rows:
            catechism_id = self._upsert_catechism(
                slug=slugify(normalize_text(row.get("Catechism ID")) or normalize_text(row.get("Reference"))),
                reference_code=normalize_text(row.get("Reference")),
                title=normalize_text(row.get("Title")) or None,
                summary=normalize_text(row.get("Summary")) or None,
            )
            for token in split_multi_value(row.get("Topics")):
                topic_id = self._resolve_topic(token)
                with self.conn.cursor() as cur:
                    cur.execute(
                        """
                        insert into public.catechism_topics (
                          catechism_reference_id, topic_id, relevance_score
                        )
                        values (%s, %s, %s)
                        on conflict (catechism_reference_id, topic_id) do nothing
                        """,
                        (catechism_id, topic_id, 5),
                    )
                self.stats.catechism_topics_upserted += 1

    def _upsert_spiritual_content_item(self, row: Dict[str, Any]) -> str:
        slug = normalize_text(row.get("Slug"))
        if not slug:
            raise ValueError("Prayer row is missing Slug.")
        title = normalize_text(row.get("Prayer Name"))
        territory_code = normalize_text(row.get("Territory")) or None
        text_status = map_text_status(row.get("Text Status"))
        is_published = text_status == "published"
        with self.conn.cursor() as cur:
            content_id = cur.execute(
                """
                insert into public.spiritual_content_items (
                  slug, title, content_kind, prayer_type, summary, body_markdown, body_html,
                  language_code, territory_code, record_type, authority_level,
                  source_label, source_url, text_status, sort_order, is_active, is_published, published_at
                )
                values (
                  %s, %s, 'prayer', %s, %s, %s, null,
                  %s, %s, %s, %s,
                  %s, %s, %s, %s, %s, %s,
                  case when %s then timezone('utc', now()) else null end
                )
                on conflict (slug) do update
                  set title = excluded.title,
                      prayer_type = excluded.prayer_type,
                      summary = coalesce(excluded.summary, public.spiritual_content_items.summary),
                      body_markdown = excluded.body_markdown,
                      language_code = excluded.language_code,
                      territory_code = excluded.territory_code,
                      record_type = excluded.record_type,
                      authority_level = excluded.authority_level,
                      source_label = excluded.source_label,
                      source_url = excluded.source_url,
                      text_status = excluded.text_status,
                      sort_order = excluded.sort_order,
                      is_active = excluded.is_active,
                      is_published = excluded.is_published,
                      published_at = case when excluded.is_published then coalesce(public.spiritual_content_items.published_at, timezone('utc', now())) else null end,
                      updated_at = timezone('utc', now())
                returning id
                """,
                (
                    slug,
                    title,
                    map_prayer_type(row.get("Prayer Type")),
                    normalize_text(row.get("Notes")) or None,
                    normalize_text(row.get("Text")) or None,
                    normalize_text(row.get("Language")) or "en",
                    territory_code,
                    normalize_text(row.get("Record Type")) or "prayer",
                    normalize_text(row.get("Authority Level")) or None,
                    normalize_text(row.get("Source Title")) or None,
                    normalize_text(row.get("Source URL")) or None,
                    text_status,
                    int(normalize_text(row.get("Sort Order")) or "0"),
                    normalize_bool(row.get("Is Active")) if normalize_text(row.get("Is Active")) else True,
                    is_published,
                    is_published,
                ),
            ).fetchone()["id"]
        self.stats.prayers_upserted += 1
        return str(content_id)

    def _replace_content_topics(self, content_id: str, tokens: Sequence[str]) -> None:
        topic_ids = []
        for token in tokens:
            try:
                topic_ids.append(self._resolve_topic(token))
            except Exception as exc:
                self.stats.warn(f"Could not resolve topic token '{token}' for content {content_id}: {exc}")

        with self.conn.cursor() as cur:
            cur.execute(
                "delete from public.spiritual_content_topics where spiritual_content_item_id = %s",
                (content_id,),
            )
            for topic_id in sorted(set(topic_ids)):
                cur.execute(
                    """
                    insert into public.spiritual_content_topics (
                      spiritual_content_item_id, topic_id, relevance_score
                    )
                    values (%s, %s, %s)
                    on conflict (spiritual_content_item_id, topic_id) do nothing
                    """,
                    (content_id, topic_id, 5),
                )
                self.stats.prayer_topics_upserted += 1

    def _resolve_group_scope_rows(self, row: Dict[str, Any]) -> List[Tuple[str, Optional[str], Optional[str]]]:
        scope = normalize_text(row.get("Scope")).lower()
        territory = normalize_text(row.get("Territory")).lower()
        group_tags = split_multi_value(row.get("Group Tags"))

        if scope in {"universal", "global"}:
            return [("global", None, None)]

        if scope == "regional":
            # Keep regional content globally visible for v1, distinguished by territory_code on the content item.
            return [("global", None, None)]

        if scope == "group_specific":
            if not group_tags:
                message = f"Prayer '{normalize_text(row.get('Prayer Name'))}' is group_specific but has no Group Tags."
                if self.allow_unmapped_group_content_as_global:
                    self.stats.warn(message + " Falling back to global.")
                    return [("global", None, None)]
                raise ValueError(message)

            resolved: List[Tuple[str, Optional[str], Optional[str]]] = []
            mapping = self.group_tag_map.get("group_tag_scopes", {}) if isinstance(self.group_tag_map, dict) else {}
            for tag in group_tags:
                tag_key = slugify(tag)
                entry = mapping.get(tag_key) or mapping.get(tag)
                if not entry:
                    if tag_key in self.org_family_id_by_code:
                        resolved.append(("organization_family", self.org_family_id_by_code[tag_key], None))
                        continue
                    if tag_key in self.local_unit_id_by_code:
                        resolved.append(("local_unit", None, self.local_unit_id_by_code[tag_key]))
                        continue
                    if self.allow_unmapped_group_content_as_global:
                        self.stats.warn(f"Prayer group tag '{tag}' was not mapped. Falling back to global.")
                        resolved.append(("global", None, None))
                        continue
                    raise ValueError(
                        f"Prayer group tag '{tag}' is not mapped. Add it to the JSON config or pass --allow-unmapped-group-content-as-global."
                    )

                scope_kind = normalize_text(entry.get("scope_kind")).lower()
                if scope_kind == "organization_family":
                    family_id = normalize_text(entry.get("organization_family_id"))
                    family_code = slugify(entry.get("organization_family_code"))
                    resolved_family_id = family_id or self.org_family_id_by_code.get(family_code)
                    if not resolved_family_id:
                        raise ValueError(f"Could not resolve organization family for group tag '{tag}'.")
                    resolved.append(("organization_family", resolved_family_id, None))
                elif scope_kind == "local_unit":
                    local_unit_id = normalize_text(entry.get("local_unit_id"))
                    local_unit_code = slugify(entry.get("local_unit_code"))
                    resolved_local_unit_id = local_unit_id or self.local_unit_id_by_code.get(local_unit_code)
                    if not resolved_local_unit_id:
                        raise ValueError(f"Could not resolve local unit for group tag '{tag}'.")
                    resolved.append(("local_unit", None, resolved_local_unit_id))
                elif scope_kind == "global":
                    resolved.append(("global", None, None))
                else:
                    raise ValueError(f"Unsupported scope_kind '{scope_kind}' for group tag '{tag}'.")

            return resolved or [("global", None, None)]

        self.stats.warn(
            f"Prayer '{normalize_text(row.get('Prayer Name'))}' has unsupported scope '{scope}'. Falling back to global."
        )
        return [("global", None, None)]

    def _replace_content_scopes(self, content_id: str, row: Dict[str, Any]) -> None:
        scope_rows = self._resolve_group_scope_rows(row)
        with self.conn.cursor() as cur:
            cur.execute(
                "delete from public.spiritual_content_scopes where spiritual_content_item_id = %s",
                (content_id,),
            )
            seen = set()
            for scope_kind, organization_family_id, local_unit_id in scope_rows:
                dedupe_key = (scope_kind, organization_family_id, local_unit_id)
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                cur.execute(
                    """
                    insert into public.spiritual_content_scopes (
                      spiritual_content_item_id, scope_kind, organization_family_id, local_unit_id
                    )
                    values (%s, %s, %s, %s)
                    """,
                    (content_id, scope_kind, organization_family_id, local_unit_id),
                )
                self.stats.prayer_scopes_replaced += 1

    def _import_prayers(self) -> None:
        rows = read_sheet_rows(self.workbook_path, "prayers")
        for row in rows:
            content_id = self._upsert_spiritual_content_item(row)
            self._replace_content_topics(content_id, split_multi_value(row.get("Topics")))
            self._replace_content_scopes(content_id, row)

    def print_summary(self) -> None:
        summary = {
            "topics_upserted": self.stats.topics_upserted,
            "topic_aliases_upserted": self.stats.topic_aliases_upserted,
            "saints_upserted": self.stats.saints_upserted,
            "saint_aliases_upserted": self.stats.saint_aliases_upserted,
            "saint_topics_upserted": self.stats.saint_topics_upserted,
            "scripture_upserted": self.stats.scripture_upserted,
            "scripture_topics_upserted": self.stats.scripture_topics_upserted,
            "catechism_upserted": self.stats.catechism_upserted,
            "catechism_topics_upserted": self.stats.catechism_topics_upserted,
            "prayers_upserted": self.stats.prayers_upserted,
            "prayer_topics_upserted": self.stats.prayer_topics_upserted,
            "prayer_scopes_replaced": self.stats.prayer_scopes_replaced,
            "skipped_patronage_categories": self.stats.skipped_patronage_categories,
            "warnings": self.stats.warnings,
        }
        print(json.dumps(summary, indent=2))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to saints_master_dataset_v1.xlsx")
    parser.add_argument(
        "--database-url",
        default=os.environ.get("DATABASE_URL", ""),
        help="Postgres connection string. Defaults to DATABASE_URL.",
    )
    parser.add_argument(
        "--group-tag-map",
        default="",
        help="Optional JSON file that maps workbook Group Tags to organization_family/local_unit scopes.",
    )
    parser.add_argument(
        "--write",
        action="store_true",
        help="Commit the import. Without this flag, the script runs in dry-run mode and rolls back.",
    )
    parser.add_argument(
        "--allow-unmapped-group-content-as-global",
        action="store_true",
        help="Fallback unmapped group_specific prayers to global scope instead of failing.",
    )
    return parser.parse_args()


def load_group_tag_map(path: str) -> Optional[Dict[str, Any]]:
    if not path:
        return None
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def main() -> None:
    args = parse_args()
    if not args.database_url:
        raise SystemExit("Missing --database-url and DATABASE_URL.")

    for sheet_name in WORKBOOK_SHEET_NAMES:
        # fail fast if the workbook shape changed
        read_sheet_rows(args.xlsx, sheet_name)

    group_tag_map = load_group_tag_map(args.group_tag_map)

    with psycopg.connect(args.database_url, row_factory=dict_row) as conn:
        importer = SpiritualImporter(
            conn=conn,
            workbook_path=args.xlsx,
            dry_run=not args.write,
            group_tag_map=group_tag_map,
            allow_unmapped_group_content_as_global=args.allow_unmapped_group_content_as_global,
        )
        try:
            importer.run()
            importer.print_summary()
            if args.write:
                conn.commit()
                print("Committed import.")
            else:
                conn.rollback()
                print("Dry run complete. Rolled back.")
        except Exception:
            conn.rollback()
            raise


if __name__ == "__main__":
    main()
